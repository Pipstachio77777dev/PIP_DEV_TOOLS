import tkinter as tk
from tkinter import filedialog, messagebox
import json
import openpyxl

class AnkiConverterApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Anki TXT to JSON/XLSX Converter")

        self.data = []
        self.field_names = []
        self.preview_labels = []
        self.field_entries = []

        # Botones
        self.load_button = tk.Button(root, text="Cargar archivo TXT", command=self.load_txt)
        self.load_button.pack(pady=5)

        self.preview_frame = tk.Frame(root)
        self.preview_frame.pack(pady=5)

        self.fields_frame = tk.Frame(root)
        self.fields_frame.pack(pady=5)

        self.define_fields_button = tk.Button(root, text="Confirmar nombres de campos", command=self.confirm_fields)
        self.define_fields_button.pack(pady=5)

        self.export_json_button = tk.Button(root, text="Exportar a JSON", command=self.export_json)
        self.export_json_button.pack(pady=5)

        self.export_excel_button = tk.Button(root, text="Exportar a Excel (XLSX)", command=self.export_excel)
        self.export_excel_button.pack(pady=5)

        self.status_label = tk.Label(root, text="", fg="blue")
        self.status_label.pack(pady=10)

    def load_txt(self):
        filepath = filedialog.askopenfilename(filetypes=[("Text files", "*.txt")])
        if not filepath:
            return

        with open(filepath, 'r', encoding='utf-8') as f:
            lines = f.readlines()

        # Filtra comentarios
        lines = [line.strip() for line in lines if line.strip() and not line.startswith("#")]

        self.data = [line.split('\t') for line in lines]

        if not self.data:
            messagebox.showerror("Error", "No se encontraron datos válidos.")
            return

        num_fields = len(self.data[0])
        self.status_label.config(text=f"{len(self.data)} filas cargadas con {num_fields} campos.")

        self.show_preview_and_field_inputs()

    def show_preview_and_field_inputs(self):
        # Limpiar vista previa anterior
        for widget in self.preview_frame.winfo_children():
            widget.destroy()
        for widget in self.fields_frame.winfo_children():
            widget.destroy()
        self.field_entries.clear()

        tk.Label(self.preview_frame, text="Vista previa de campos (primeras 3 filas):", font=('Arial', 10, 'bold')).pack()

        table_frame = tk.Frame(self.preview_frame)
        table_frame.pack()

        preview_rows = self.data[:3]

        # Mostrar encabezados por defecto
        for col in range(len(self.data[0])):
            tk.Label(table_frame, text=f"Campo {col + 1}", font=('Arial', 9, 'bold'), borderwidth=1, relief='solid', padx=5, pady=2).grid(row=0, column=col, sticky='nsew')

        # Mostrar datos
        for row_index, row in enumerate(preview_rows):
            for col_index, value in enumerate(row):
                tk.Label(table_frame, text=value, borderwidth=1, relief='solid', padx=5, pady=2).grid(row=row_index + 1, column=col_index, sticky='nsew')

        # Campos para asignar nombres
        tk.Label(self.fields_frame, text="Asignar nombres a los campos:", font=('Arial', 10, 'bold')).pack()
        for i in range(len(self.data[0])):
            frame = tk.Frame(self.fields_frame)
            frame.pack(pady=2)
            tk.Label(frame, text=f"Campo {i + 1}:", width=10, anchor='w').pack(side='left')
            entry = tk.Entry(frame, width=30)
            entry.insert(0, f"campo_{i + 1}")
            entry.pack(side='left')
            self.field_entries.append(entry)

    def confirm_fields(self):
        if not self.field_entries:
            messagebox.showerror("Error", "Primero debes cargar un archivo y definir los campos.")
            return

        self.field_names = [entry.get().strip() for entry in self.field_entries]

        if any(not name for name in self.field_names):
            messagebox.showerror("Error", "Todos los campos deben tener nombre.")
            return

        self.status_label.config(text=f"Campos definidos: {', '.join(self.field_names)}")

    def export_json(self):
        if not self.data or not self.field_names:
            messagebox.showerror("Error", "Carga datos y define los campos primero.")
            return

        export_data = [dict(zip(self.field_names, row)) for row in self.data]

        filepath = filedialog.asksaveasfilename(defaultextension=".json", filetypes=[("JSON files", "*.json")])
        if not filepath:
            return

        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(export_data, f, ensure_ascii=False, indent=4)

        messagebox.showinfo("Éxito", f"Archivo JSON exportado a:\n{filepath}")

    def export_excel(self):
        if not self.data or not self.field_names:
            messagebox.showerror("Error", "Carga datos y define los campos primero.")
            return

        filepath = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if not filepath:
            return

        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # Encabezados
        for col_index, name in enumerate(self.field_names, start=1):
            sheet.cell(row=1, column=col_index, value=name)

        # Datos
        for row_index, row in enumerate(self.data, start=2):
            for col_index, value in enumerate(row, start=1):
                sheet.cell(row=row_index, column=col_index, value=value)

        workbook.save(filepath)
        messagebox.showinfo("Éxito", f"Archivo XLSX exportado a:\n{filepath}")

if __name__ == "__main__":
    root = tk.Tk()
    app = AnkiConverterApp(root)
    root.mainloop()
